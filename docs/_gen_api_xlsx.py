"""Generate Backend-APIs.xlsx from the current Go router catalogue."""
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "Backend-APIs.xlsx"

thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
title_font = Font(name="Calibri", bold=True, color="1E3A5F", size=16)
subtitle_font = Font(name="Calibri", size=11, color="475467")
body_font = Font(name="Calibri", size=10)
wrap = Alignment(wrap_text=True, vertical="center")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")
alt_fill = PatternFill("solid", fgColor="F8FAFC")
note_fill = PatternFill("solid", fgColor="FFF4E5")
unused_fill = PatternFill("solid", fgColor="FEF3F2")
yes_fill = PatternFill("solid", fgColor="D1FADF")
no_fill = PatternFill("solid", fgColor="F2F4F7")

method_fills = {
    "GET": PatternFill("solid", fgColor="D1FADF"),
    "POST": PatternFill("solid", fgColor="D1E9FF"),
    "PUT": PatternFill("solid", fgColor="FEF0C7"),
    "PATCH": PatternFill("solid", fgColor="E0EAFF"),
    "DELETE": PatternFill("solid", fgColor="FEE4E2"),
}
method_fonts = {
    "GET": Font(name="Calibri", bold=True, color="027A48", size=10),
    "POST": Font(name="Calibri", bold=True, color="175CD3", size=10),
    "PUT": Font(name="Calibri", bold=True, color="B54708", size=10),
    "PATCH": Font(name="Calibri", bold=True, color="3538CD", size=10),
    "DELETE": Font(name="Calibri", bold=True, color="B42318", size=10),
}

# module, method, path, auth, role, description, params, body, status, used
APIS = [
    (
        "Health", "GET", "/health", "None", "Public",
        "Liveness check. Returns {status: ok}.",
        "—", "—", 200, "No (ops only)",
    ),
    (
        "Public", "GET", "/api/halls", "None", "Public",
        "List all active seminar halls.",
        "—", "—", 200, "Yes",
    ),
    (
        "Public", "GET", "/api/halls/:id", "None", "Public",
        "Get a single hall by ID.",
        "id (path, int) — hall_id", "—", 200, "Yes (hall calendar)",
    ),
    (
        "Public", "GET", "/api/availability", "None", "Public",
        "Hall availability. Omit hall_id for all halls on one date; pass hall_id + date for one hall; or hall_id + start_date + end_date for a range.",
        "date (YYYY-MM-DD) OR start_date + end_date; hall_id optional",
        "—", 200, "Yes (home, dashboard, calendar)",
    ),
    (
        "Public", "GET", "/api/departments", "None", "Public",
        "List all departments.",
        "—", "—", 200, "Yes",
    ),
    (
        "Public", "GET", "/api/clubs", "None", "Public",
        "List clubs. Optionally filter by department.",
        "dept_id (query, optional)", "—", 200, "Yes",
    ),
    (
        "Auth", "POST", "/api/auth/login", "None", "Public",
        "Login. Returns JWT and user. Also sets HTTP-only cookie auth_token.",
        "—", '{"username":"admin","password":"admin123"}', 200, "Yes",
    ),
    (
        "Auth", "POST", "/api/auth/logout", "None", "Public",
        "Clears auth_token cookie.",
        "—", "—", 200, "Yes",
    ),
    (
        "Auth", "GET", "/api/auth/me", "JWT", "Any authenticated",
        "Current user from token. Requesters may also get a requester object.",
        "—", "—", 200, "Yes",
    ),
    (
        "Auth", "PATCH", "/api/auth/password", "JWT", "Any authenticated",
        "Change own password. new_password min 6 chars. JWT is not rotated.",
        "—", '{"current_password":"...","new_password":"..."}', 200, "Yes",
    ),
    (
        "Requester", "POST", "/api/requests", "JWT", "requester",
        "Create a booking request (class/club representative).",
        "—",
        '{"hall_id":1,"event_title":"...","event_description":"...","event_date":"YYYY-MM-DD","start_time":"HH:MM","end_time":"HH:MM","expected_attendees":80,"purpose":"..."}',
        201, "Yes",
    ),
    (
        "Requester", "GET", "/api/requests/my", "JWT", "requester",
        "List booking requests created by the logged-in requester.",
        "—", "—", 200, "Yes",
    ),
    (
        "Dept Coordinator", "GET", "/api/requests/dept-pending", "JWT", "dept_coordinator",
        "Pending requests for this coordinator's department.",
        "—", "—", 200, "Yes",
    ),
    (
        "Dept Coordinator", "PATCH", "/api/requests/:id/dept-review", "JWT", "dept_coordinator",
        "Approve or reject a request at department level. action: approve | reject.",
        "id (path, int) — request_id", '{"action":"approve","remarks":"..."}', 200, "Yes",
    ),
    (
        "Dept Coordinator", "GET", "/api/users/faculties", "JWT", "dept_coordinator",
        "List users with role faculty. Not scoped to coordinator's department.",
        "—", "—", 200, "No (unused; UI link removed)",
    ),
    (
        "Classes", "GET", "/api/classes", "JWT", "admin or dept_coordinator",
        "List classes for a department.",
        "dept_id (query, required)", "—", 200, "Yes (admin classes/users pages)",
    ),
    (
        "Classes", "POST", "/api/classes", "JWT", "admin or dept_coordinator",
        "Create a class.",
        "—", '{"class_name":"BCA Semester 3","dept_id":1,"year":"2025-26"}', 201, "Yes (admin classes page)",
    ),
    (
        "Classes", "DELETE", "/api/classes/:id", "JWT", "admin or dept_coordinator",
        "Delete a class.",
        "id (path, int) — class_id", "—", 200, "Yes (admin classes page)",
    ),
    (
        "Admin — Requests", "GET", "/api/requests/admin-pending", "JWT", "admin",
        "Requests awaiting final admin approval.",
        "—", "—", 200, "Yes",
    ),
    (
        "Admin — Requests", "PATCH", "/api/requests/:id/admin-review", "JWT", "admin",
        "Final approve or reject. Approve creates a booking (409 if slot overlap).",
        "id (path, int) — request_id", '{"action":"approve","remarks":"..."}', 200, "Yes",
    ),
    (
        "Admin — Bookings", "GET", "/api/bookings", "JWT", "admin",
        "List all bookings.",
        "—", "—", 200, "Yes",
    ),
    (
        "Admin — Bookings", "PATCH", "/api/bookings/:id/cancel", "JWT", "admin",
        "Cancel a booking (fails if already cancelled).",
        "id (path, int) — booking_id", "—", 200, "Yes",
    ),
    (
        "Admin — Departments", "POST", "/api/departments", "JWT", "admin",
        "Create a department.",
        "—", '{"dept_name":"Electronics","dept_code":"ECE"}', 201, "Yes",
    ),
    (
        "Admin — Departments", "PUT", "/api/departments/:id", "JWT", "admin",
        "Update a department.",
        "id (path, int) — dept_id", '{"dept_name":"Computer Science","dept_code":"CS"}',
        200, "No (wrapper exists; no edit form)",
    ),
    (
        "Admin — Departments", "DELETE", "/api/departments/:id", "JWT", "admin",
        "Delete a department.",
        "id (path, int) — dept_id", "—", 200, "Yes",
    ),
    (
        "Admin — Halls", "POST", "/api/halls", "JWT", "admin",
        "Create a hall. facilities is an arbitrary JSON object.",
        "—",
        '{"hall_name":"Main Auditorium","capacity":300,"location":"Admin Block","facilities":{"projector":true}}',
        201, "Yes",
    ),
    (
        "Admin — Halls", "PUT", "/api/halls/:id", "JWT", "admin",
        "Update a hall including is_active.",
        "id (path, int) — hall_id",
        '{"hall_name":"...","capacity":300,"location":"...","facilities":{},"is_active":true}',
        200, "Yes",
    ),
    (
        "Admin — Halls", "DELETE", "/api/halls/:id", "JWT", "admin",
        "Soft-deactivate a hall (is_active = false).",
        "id (path, int) — hall_id", "—", 200, "Yes",
    ),
    (
        "Admin — Users", "GET", "/api/users", "JWT", "admin",
        "List users including inactive. Optional role filter.",
        "role (query, optional): admin | dept_coordinator | requester | faculty",
        "—", 200, "Yes",
    ),
    (
        "Admin — Users", "POST", "/api/users", "JWT", "admin",
        "Create a user. For role requester: dept_id and requester_type (class|club) required; class_id or club_id as applicable.",
        "—",
        '{"username":"student02","password":"req1234","full_name":"...","email":"...","phone":"...","role":"requester","dept_id":1,"requester_type":"class","class_id":1}',
        201, "Yes",
    ),
    (
        "Admin — Users", "DELETE", "/api/users/:id", "JWT", "admin",
        "Deactivate a user (is_active = false).",
        "id (path, int) — user_id", "—", 200, "Yes",
    ),
    (
        "Admin — Users", "PATCH", "/api/users/:id/reactivate", "JWT", "admin",
        "Reactivate a previously deactivated user.",
        "id (path, int) — user_id", "—", 200, "Yes",
    ),
    (
        "Admin — Users", "PATCH", "/api/users/:id/password", "JWT", "admin",
        "Admin sets another user's password. new_password min 6 chars. Does not require old password.",
        "id (path, int) — user_id", '{"new_password":"newpass123"}', 200, "Yes",
    ),
    (
        "Admin — Clubs", "POST", "/api/clubs", "JWT", "admin",
        "Create a club.",
        "—", '{"club_name":"Tech Club","dept_id":1,"description":"..."}', 201, "Yes",
    ),
    (
        "Admin — Clubs", "DELETE", "/api/clubs/:id", "JWT", "admin",
        "Delete a club.",
        "id (path, int) — club_id", "—", 200, "Yes",
    ),
    (
        "Admin — Reports", "GET", "/api/reports/hall-usage", "JWT", "admin",
        "Hall usage report (booking counts per hall).",
        "—", "—", 200, "Yes",
    ),
    (
        "Admin — Reports", "GET", "/api/reports/department-usage", "JWT", "admin",
        "Department usage report (request counts per department).",
        "—", "—", 200, "Yes",
    ),
]


def apply_header_row(ws, row, headers):
    for col, heading in enumerate(headers, 1):
        cell = ws.cell(row, col, heading)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin


def ticks(auth, role_field):
    if auth == "None":
        return ("✓", "✓", "✓", "✓", "✓")
    if "Any authenticated" in role_field:
        return ("—", "✓", "✓", "✓", "✓")
    if role_field == "requester":
        return ("—", "✓", "—", "—", "—")
    if role_field == "dept_coordinator":
        return ("—", "—", "✓", "—", "—")
    if "admin or dept_coordinator" in role_field:
        return ("—", "—", "✓", "✓", "—")
    if role_field == "admin":
        return ("—", "—", "—", "✓", "—")
    return ("—", "—", "—", "—", "—")


def build():
    wb = Workbook()

    # ── Sheet: All APIs ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All APIs"
    ws.sheet_properties.tabColor = "1E3A5F"

    ws.merge_cells("A1:K1")
    ws["A1"] = "Seminar Hall Booking System — Backend API Catalogue"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        "Source: backend/internal/router/router.go  •  "
        "Base URL: http://localhost:8080  •  "
        "Auth: Authorization: Bearer <JWT> (or cookie auth_token)  •  "
        f"Total endpoints: {len(APIS)}"
    )
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8

    headers = [
        "S.No", "Module", "Method", "Endpoint", "Auth", "Role",
        "Description", "Path / Query params", "Request body (JSON)",
        "Success status", "Used by frontend",
    ]
    apply_header_row(ws, 4, headers)
    ws.row_dimensions[4].height = 22
    ws.auto_filter.ref = f"A4:K{4 + len(APIS)}"
    ws.freeze_panes = "A5"

    for i, row in enumerate(APIS, 1):
        r = 4 + i
        module, method, path, auth, role, desc, params, body, status, used = row
        values = [i, module, method, path, auth, role, desc, params, body, status, used]
        for c, val in enumerate(values, 1):
            cell = ws.cell(r, c, val)
            cell.border = thin
            cell.alignment = center if c in (1, 3, 5, 10) else wrap
            cell.font = body_font
            if i % 2 == 0:
                cell.fill = alt_fill
        mcell = ws.cell(r, 3)
        mcell.fill = method_fills[method]
        mcell.font = method_fonts[method]
        mcell.alignment = center
        if str(used).startswith("No"):
            ws.cell(r, 11).fill = unused_fill
            ws.cell(r, 11).font = Font(name="Calibri", size=10, color="B42318")
        ws.row_dimensions[r].height = 48 if len(desc) > 80 or (body and body != "—") else 32

    widths = {
        "A": 7, "B": 22, "C": 10, "D": 38, "E": 10, "F": 24,
        "G": 48, "H": 42, "I": 55, "J": 14, "K": 28,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = "1:4"
    ws.sheet_view.showGridLines = False
    ws.oddHeader.left.text = "Seminar Hall Booking System"
    ws.oddHeader.right.text = "Backend APIs"
    ws.oddFooter.center.text = "Page &P of &N"

    # ── Sheet: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary", 0)
    ws2.sheet_properties.tabColor = "175CD3"
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "API summary by module and method"
    ws2["A1"].font = title_font
    ws2.merge_cells("A2:D2")
    ws2["A2"] = "Generated from backend/internal/router/router.go"
    ws2["A2"].font = subtitle_font

    mod_counts = Counter(a[0] for a in APIS)
    meth_counts = Counter(a[1] for a in APIS)
    used_yes = sum(1 for a in APIS if str(a[9]).startswith("Yes"))
    used_no = len(APIS) - used_yes

    ws2["A4"] = "By module"
    ws2["A4"].font = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
    apply_header_row(ws2, 5, ["Module", "Count"])
    for i, (mod, n) in enumerate(mod_counts.items(), 1):
        ws2.cell(5 + i, 1, mod).border = thin
        ws2.cell(5 + i, 2, n).border = thin
        ws2.cell(5 + i, 2).alignment = center
        if i % 2 == 0:
            ws2.cell(5 + i, 1).fill = alt_fill
            ws2.cell(5 + i, 2).fill = alt_fill
    mod_end = 5 + len(mod_counts)
    ws2.cell(mod_end + 1, 1, "Total").font = Font(bold=True)
    ws2.cell(mod_end + 1, 2, len(APIS)).font = Font(bold=True)
    ws2.cell(mod_end + 1, 1).border = thin
    ws2.cell(mod_end + 1, 2).border = thin

    meth_start = mod_end + 4
    ws2.cell(meth_start, 1, "By HTTP method").font = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
    apply_header_row(ws2, meth_start + 1, ["Method", "Count"])
    for i, method in enumerate(["GET", "POST", "PUT", "PATCH", "DELETE"], 1):
        ws2.cell(meth_start + 1 + i, 1, method).border = thin
        ws2.cell(meth_start + 1 + i, 1).fill = method_fills[method]
        ws2.cell(meth_start + 1 + i, 1).font = method_fonts[method]
        ws2.cell(meth_start + 1 + i, 2, meth_counts.get(method, 0)).border = thin
        ws2.cell(meth_start + 1 + i, 2).alignment = center

    used_start = meth_start + 9
    ws2.cell(used_start, 1, "Frontend usage").font = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
    apply_header_row(ws2, used_start + 1, ["Status", "Count"])
    ws2.cell(used_start + 2, 1, "Called by frontend")
    ws2.cell(used_start + 2, 2, used_yes)
    ws2.cell(used_start + 3, 1, "Not called by frontend")
    ws2.cell(used_start + 3, 2, used_no)
    for r in range(used_start + 2, used_start + 4):
        ws2.cell(r, 1).border = thin
        ws2.cell(r, 2).border = thin
        ws2.cell(r, 2).alignment = center
    ws2.cell(used_start + 3, 1).fill = unused_fill

    ws2.cell(used_start + 6, 1, "Not called by the current frontend").font = Font(
        name="Calibri", bold=True, size=12, color="1E3A5F"
    )
    ws2.merge_cells(f"A{used_start + 6}:C{used_start + 6}")
    apply_header_row(ws2, used_start + 7, ["Method", "Endpoint", "Notes"])
    unused_rows = [
        ("GET", "/health", "Health check; not used by the UI"),
        ("GET", "/api/users/faculties", "Coordinator faculty list; UI link removed"),
        ("PUT", "/api/departments/:id", "departmentsApi.update exists; no edit form"),
    ]
    for i, (m, p, n) in enumerate(unused_rows, 1):
        r = used_start + 7 + i
        ws2.cell(r, 1, m).fill = method_fills[m]
        ws2.cell(r, 1).font = method_fonts[m]
        ws2.cell(r, 1).alignment = center
        ws2.cell(r, 2, p)
        ws2.cell(r, 3, n)
        for c in range(1, 4):
            ws2.cell(r, c).border = thin

    note_row = used_start + 12
    ws2.merge_cells(f"A{note_row}:C{note_row + 3}")
    ws2[f"A{note_row}"] = (
        "Notes\n"
        "• JWT: send header Authorization: Bearer <token> (login also sets cookie auth_token).\n"
        "• Seed logins: admin / admin123  •  coordinator1 / coord123  •  student1 / req123\n"
        "• GET /api/users/faculties is still registered in the router but is unused by the UI.\n"
        "• Faculty is a user role (users.role enum), not a separate table."
    )
    ws2[f"A{note_row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws2[f"A{note_row}"].fill = note_fill
    ws2.row_dimensions[note_row].height = 72

    chart = PieChart()
    chart.title = "APIs by HTTP method"
    labels = Reference(ws2, min_col=1, min_row=meth_start + 2, max_row=meth_start + 6)
    data = Reference(ws2, min_col=2, min_row=meth_start + 1, max_row=meth_start + 6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.width = 12
    chart.height = 8
    ws2.add_chart(chart, "E4")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 48
    ws2.column_dimensions["D"].width = 16
    ws2.row_dimensions[1].height = 24
    ws2.sheet_view.showGridLines = False

    # ── Sheet: Auth & roles ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Auth & roles")
    ws3.sheet_properties.tabColor = "027A48"
    ws3.merge_cells("A1:G1")
    ws3["A1"] = "Who can call which API"
    ws3["A1"].font = title_font
    ws3.merge_cells("A2:G2")
    ws3["A2"] = "✓ = allowed  •  — = not allowed (401/403)"
    ws3["A2"].font = subtitle_font

    apply_header_row(
        ws3, 4,
        ["Endpoint", "Method", "Public", "requester", "dept_coordinator", "admin", "faculty"],
    )
    for i, a in enumerate(APIS, 1):
        r = 4 + i
        _module, method, path, auth, role = a[0], a[1], a[2], a[3], a[4]
        ws3.cell(r, 1, path).border = thin
        ws3.cell(r, 2, method).border = thin
        ws3.cell(r, 2).fill = method_fills[method]
        ws3.cell(r, 2).font = method_fonts[method]
        ws3.cell(r, 2).alignment = center
        for c, t in enumerate(ticks(auth, role), 3):
            cell = ws3.cell(r, c, t)
            cell.alignment = center
            cell.border = thin
            cell.fill = yes_fill if t == "✓" else no_fill
        if i % 2 == 0:
            ws3.cell(r, 1).fill = alt_fill

    ws3.freeze_panes = "A5"
    ws3.auto_filter.ref = f"A4:G{4 + len(APIS)}"
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 10
    for col in "CDEFG":
        ws3.column_dimensions[col].width = 18
    ws3.sheet_view.showGridLines = False
    ws3.page_setup.orientation = "landscape"
    ws3.page_setup.fitToPage = True
    ws3.page_setup.fitToWidth = 1
    ws3.page_setup.fitToHeight = 0
    ws3.page_setup.paperSize = ws3.PAPERSIZE_A4

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}  sheets={wb.sheetnames}  endpoints={len(APIS)}")


if __name__ == "__main__":
    build()
